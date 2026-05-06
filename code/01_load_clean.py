"""
01_load_clean.py
Parse Data_final.xlsm (Bloomberg pull) into a tidy long-form panel + benchmark CSV.

Inputs : ../Data_final.xlsm
Outputs: ../data/processed/panel.csv      (date, ticker, price, sentiment, mkt_cap, log_ret)
         ../data/processed/benchmarks.csv (date, spx, vix, spx_ret)
"""
from pathlib import Path
import numpy as np
import pandas as pd
import openpyxl

PROJECT = Path(__file__).resolve().parent.parent
RAW     = PROJECT / "Data_final.xlsm"
OUT_DIR = PROJECT / "data" / "processed"
OUT_DIR.mkdir(parents=True, exist_ok=True)

TICKERS = [
    "AAPL", "MSFT", "GOOGL", "NVDA", "META", "AMZN", "TSLA",
    "JPM",  "BAC",  "GS",    "WFC",
    "JNJ",  "UNH",  "PFE",   "LLY",  "MRK",
    "WMT",  "PG",   "KO",    "PEP",  "HD",   "NKE",  "DIS",
    "XOM",  "CVX",
    "BA",   "CAT",  "GE",
    "T",    "VZ",
]

SECTOR = {
    "AAPL":"Tech","MSFT":"Tech","GOOGL":"Tech","NVDA":"Tech","META":"Tech",
    "AMZN":"ConsDisc","TSLA":"ConsDisc","HD":"ConsDisc","NKE":"ConsDisc","DIS":"CommSvc",
    "T":"CommSvc","VZ":"CommSvc",
    "JPM":"Fin","BAC":"Fin","GS":"Fin","WFC":"Fin",
    "JNJ":"Hlth","UNH":"Hlth","PFE":"Hlth","LLY":"Hlth","MRK":"Hlth",
    "WMT":"Stap","PG":"Stap","KO":"Stap","PEP":"Stap",
    "XOM":"Engy","CVX":"Engy",
    "BA":"Indu","CAT":"Indu","GE":"Indu",
}


def _num(v):
    """Coerce strings like 'N/A' / '#N/A...' to NaN; pass through real numbers."""
    if isinstance(v, (int, float)) and not (isinstance(v, bool)):
        return float(v)
    return np.nan


def parse_ticker_sheet(wb, ticker):
    """Extract (date, price, sentiment, mkt_cap) from a ticker sheet.

    Layout in each sheet (row 1 is headers):
      col 0 = Price Date     col 1 = Price
      col 3 = Sentiment Date col 4 = Sentiment
      col 6 = Stories Date   col 7 = NumStories  (broken — Bloomberg returns N/A)
      col 9 = MktCap Date    col 10 = MktCap
    """
    ws = wb[ticker]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        records.append({
            "date":      row[0],
            "ticker":    ticker,
            "price":     _num(row[1]) if len(row) > 1  else np.nan,
            "sentiment": _num(row[4]) if len(row) > 4  else np.nan,
            "mkt_cap":   _num(row[10]) if len(row) > 10 else np.nan,
        })
    return pd.DataFrame(records)


def parse_benchmarks(wb):
    ws = wb["BENCHMARKS"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        spx_date = row[0] if len(row) > 0 else None
        spx_val  = _num(row[1]) if len(row) > 1 else np.nan
        vix_date = row[3] if len(row) > 3 else None
        vix_val  = _num(row[4]) if len(row) > 4 else np.nan
        if spx_date is not None:
            rows.append({"date": spx_date, "series": "spx", "value": spx_val})
        if vix_date is not None:
            rows.append({"date": vix_date, "series": "vix", "value": vix_val})
    long = pd.DataFrame(rows)
    wide = long.pivot_table(index="date", columns="series", values="value").reset_index()
    wide["date"] = pd.to_datetime(wide["date"])
    wide = wide.sort_values("date").reset_index(drop=True)
    wide["spx_ret"] = wide["spx"].pct_change()
    return wide


def main():
    print(f"Reading {RAW.name} ...")
    wb = openpyxl.load_workbook(RAW, data_only=True, read_only=True)
    sheets = set(wb.sheetnames)

    panels = []
    for t in TICKERS:
        if t not in sheets:
            print(f"  WARN: sheet {t} missing — skipped")
            continue
        df = parse_ticker_sheet(wb, t)
        df["sector"] = SECTOR.get(t, "Other")
        panels.append(df)
        print(f"  {t:6s}  rows={len(df):5d}  "
              f"price_nonnull={df['price'].notna().sum():5d}  "
              f"sent_nonnull={df['sentiment'].notna().sum():5d}  "
              f"mcap_nonnull={df['mkt_cap'].notna().sum():5d}")

    panel = pd.concat(panels, ignore_index=True)
    panel["date"] = pd.to_datetime(panel["date"])
    panel = panel.sort_values(["ticker", "date"]).reset_index(drop=True)

    panel = panel[panel["price"].notna()].copy()
    panel["log_ret"] = (
        panel.groupby("ticker")["price"]
             .transform(lambda s: np.log(s).diff())
    )

    panel_out = OUT_DIR / "panel.csv"
    panel.to_csv(panel_out, index=False)
    print(f"\nWrote {panel_out}")
    print(f"  rows = {len(panel):,}")
    print(f"  tickers = {panel['ticker'].nunique()}")
    print(f"  date range = {panel['date'].min().date()} -> {panel['date'].max().date()}")

    bench = parse_benchmarks(wb)
    bench_out = OUT_DIR / "benchmarks.csv"
    bench.to_csv(bench_out, index=False)
    print(f"\nWrote {bench_out}")
    print(f"  rows = {len(bench):,}")
    print(f"  date range = {bench['date'].min().date()} -> {bench['date'].max().date()}")
    print(f"  SPX range: {bench['spx'].min():.0f} -> {bench['spx'].max():.0f}")
    print(f"  VIX range: {bench['vix'].min():.2f} -> {bench['vix'].max():.2f}")

    print("\nDone.")


if __name__ == "__main__":
    main()
